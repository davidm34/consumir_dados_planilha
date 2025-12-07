import pandas as pd
from openpyxl import load_workbook

arquivo_origem = "CAETITE_INDICADORES CONTROLE MENSAL GECEN - 2024 - Macro Reservatório (1).xlsx"
arquivo_destino = "TRATAMENTO DE DADOS.xlsx"
nome_aba_macro = 'LEITURA VOL. MACRO_m³'
nome_aba_micro = 'LEITURA VOL. MICRO_m³'
nome_aba_dados = 'DIMENSÃO'

def gerar_tabela_dados():
    print("1. Lendo dados do Excel...")
    try:
        # Lê sem cabeçalho para usar índices numéricos
        df_macro = pd.read_excel(arquivo_origem, sheet_name=nome_aba_macro, header=None)
        df_micro = pd.read_excel(arquivo_origem, sheet_name=nome_aba_micro, header=None)
    except FileNotFoundError:
        print(f"ERRO: Arquivo '{arquivo_origem}' não encontrado.")
        return

    linha = 10
    col_sistema = 1       # Coluna B
    col_lig_totais = 39   # Coluna AN
    col_lig_ativas = 28   # Coluna AC

    sistemas = []
    totais = []
    ativas = []

    # Loop de Leitura
    while True:
        if linha >= len(df_macro) or linha >= len(df_micro):
            break

        nome = df_macro.iloc[linha, col_sistema]

        # Condição de parada
        if pd.isna(nome) or str(nome).strip().upper().startswith('TOTAL'):
            break

        val_total = df_macro.iloc[linha, col_lig_totais]
        val_ativa = df_micro.iloc[linha, col_lig_ativas]

        # Tratamento de número
        def limpar(v):
            try:
                return int(float(str(v).strip()))
            except:
                return 0

        # Só adiciona se o sistema tiver nome
        if str(nome).strip() != "":
            sistemas.append(str(nome).strip())
            totais.append(limpar(val_total))
            ativas.append(limpar(val_ativa))

        linha += 1
    
    print(f"   -> {len(sistemas)} sistemas processados.")

    # 2. INSERINDO NO EXCEL
    print("2. Inserindo a tabela no Excel...")
    try:
        book = load_workbook(arquivo_destino)
    except Exception as e:
        print(f"Erro ao abrir Excel de destino: {e}")
        return

    # Prepara a aba (deleta se existir para criar limpa)
    if nome_aba_dados in book.sheetnames:
        del book[nome_aba_dados]
    ws = book.create_sheet(nome_aba_dados)

    # Escreve o cabeçalho
    ws.append(['SISTEMA', 'LIGAÇÕES TOTAIS', 'LIGAÇÕES ATIVAS'])
    
    # Escreve os dados linha por linha
    for i in range(len(sistemas)):
        ws.append([sistemas[i], totais[i], ativas[i]])

    try:
        book.save(arquivo_destino)
        print(f"\nSUCESSO! Tabela de dados inserida na aba '{nome_aba_dados}'.")
    except PermissionError:
        print("\nERRO CRÍTICO: O Excel de destino está aberto. Feche o arquivo e tente novamente.")

if __name__ == "__main__":
    gerar_tabela_dados()